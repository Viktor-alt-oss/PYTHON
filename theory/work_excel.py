
import openpyxl
import json

book = openpyxl.open(r"C:\Users\virgi\OneDrive\Рабочий стол\Единобоства\200.xlsx", read_only=True)

sheet = book.active
sheet_2 = book.worksheets[2]
print(sheet['A2'].value)
print(sheet['B2'].value)
for row in range(1, sheet.max_row + 1):
    a = sheet[row][0].value
    b = sheet[row][1].value
    c = sheet[row][2].value
    print(a, b, c)

cells = sheet['B1' : 'C11']
for name, group in cells:
    print(name.value, group.value)

for row in sheet.iter_rows(min_row=2, max_row=20, min_col=1, max_col=3):
    for cell in row:
        print(cell.value, end=' ')
    print()

with open(r"C:\Users\virgi\work\Python\gfgd.json") as file:
    data = json.load(file)
for i in data['movies']:
    id = i['id']
    title = i['title']
    print(id, title)

book1 = openpyxl.Workbook()
sheet1 = book1.active

sheet1['A1'] = 'ID'
sheet1['B1'] = 'TITLE'
sheet1['C1'] ='GENRES'

row = 2
for movie in data['movies']:
    sheet1[row][0].value = movie['id']
    sheet1[row][1].value = movie['title']
    sheet1[row][2].value = ' '.join(movie['genres'])
    row += 1


book1.save('my_book.xlsx')
book1.close()